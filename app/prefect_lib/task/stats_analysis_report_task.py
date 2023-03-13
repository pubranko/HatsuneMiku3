from copy import deepcopy
import os
import sys
from typing import Any, Final, Optional
import pandas as pd
from pydantic import ValidationError
from prefect.engine import state
from prefect.engine.runner import ENDRUN
from pymongo.cursor import Cursor
from decimal import Decimal, ROUND_HALF_UP
# from openpyxl import Workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.workbook.child import _WorkbookChild

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
from openpyxl.chart.bar_chart import BarChart
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter
path = os.getcwd()
sys.path.append(path)
from shared.settings import DATA_DIR
from shared.timezone_recovery import timezone_recovery
from BrownieAtelierNotice.mail_attach_send import mail_attach_send
from BrownieAtelierMongo.collection_models.stats_info_collect_model import StatsInfoCollectModel
from prefect_lib.task.extentions_task import ExtensionsTask
from prefect_lib.data_models.stats_analysis_report_input import StatsAnalysisReportInput
from prefect_lib.data_models.stats_info_collect_data import StatsInfoCollectData


class StatsAnalysisReportTask(ExtensionsTask):
    '''あとで'''

    # 作成するレポートの見出し項目
    HEAD1: Final[str] = 'head1'
    HEAD2: Final[str] = 'head2'
    COL: Final[str] = 'col'
    EQUIVALENT_COLOR: Final[str] = 'equivalent_color'
    WARNING_VALUE: Final[str] = 'warning_value'
    WARNING_BACKGROUND_COLOR: Final[str] = 'warning_background_color'
    DIGIT_ADJUSTMENT: Final[str] = 'digit_adjustment'
    NUMBER_FORMAT: Final[str] = 'number_format'
    WARNING_VALUE_OVER: Final[str] = 'warning_value_over'

    robots_analysis_columns_info: list = [
        # head1                     : 必須 : 見出し１行目
        # col                       : 必須 : データフレーム列名
        # equivalent_color          : 任意 : 上のセルと同値の場合の文字色。上と同値の場合、文字色を薄くするなどに使用する。
        # warning_value             : 任意 : ワーニングとする値を入れる配列
        # warning_background_color  : 任意 : ワーンングとなったセルの背景色
        {HEAD1: '集計日〜', COL: 'aggregate_base_term'},
        {HEAD1: 'スパイダー名', COL: 'spider_name',
         EQUIVALENT_COLOR: 'bbbbbb'},
        {HEAD1: 'status', COL: 'robots_response_status',
         WARNING_VALUE: [], WARNING_BACKGROUND_COLOR:'FF8C00'},
        {HEAD1: 'count', COL: 'count'},
    ]

    downloader_analysis_columns_info: list = [
        # head1                     : 必須 : 見出し１行目
        # col                       : 必須 : データフレーム列名
        # equivalent_color          : 任意 : 上のセルと同値の場合の文字色。上と同値の場合、文字色を薄くするなどに使用する。
        # warning_value             : 任意 : ワーニングとする値を入れる配列
        # warning_background_color  : 任意 : ワーンングとなったセルの背景色
        {HEAD1: '集計日〜', COL: 'aggregate_base_term'},
        {HEAD1: 'スパイダー名', COL: 'spider_name',
         EQUIVALENT_COLOR: 'bbbbbb'},
        {HEAD1: 'status', COL: 'downloader_response_status',
         WARNING_VALUE: [], WARNING_BACKGROUND_COLOR:'FF8C00'},
        {HEAD1: 'count', COL: 'count'},
    ]

    stats_analysis_columns_info: list = [
        # head1                     : 必須 : 見出し１行目
        # head2                     : 必須 : 見出し２行目
        # col                       : 必須 : データフレーム列名
        # digit_adjustment          : 任意 : 単位の調整。1000とした場合、'value/1000'となる。
        # number_format             : 任意 : 小数点以下の桁数。省略した場合'#,##0'
        # equivalent_color          : 任意 : 上のセルと同値の場合の文字色。上と同値の場合、文字色を薄くするなどに使用する。
        # warning_value             : 任意 : ワーニングとする値を入れる配列
        # warning_value_over        : 任意 : 超過したらワーニングとする値
        # warning_background_color  : 任意 : ワーンングとなったセルの背景色
        {HEAD1: '集計日〜', HEAD2: '', COL: 'aggregate_base_term'},
        {HEAD1: 'スパイダー名', HEAD2: '', COL: 'spider_name',
         EQUIVALENT_COLOR: 'bbbbbb'},
        {HEAD1: 'ログレベル件数', HEAD2: 'CRITICAL', COL: 'log_count/CRITICAL',
         WARNING_VALUE_OVER: 0, WARNING_BACKGROUND_COLOR: 'FF8C00'},    # 0件を超えた場合ワーニング
        {HEAD1: '', HEAD2: 'ERROR', COL: 'log_count/ERROR',
         WARNING_VALUE_OVER: 0, WARNING_BACKGROUND_COLOR: 'FF8C00'},    # 0件を超えた場合ワーニング
        {HEAD1: '', HEAD2: 'WARNING', COL: 'log_count/WARNING'},
        {HEAD1: '処理時間(秒)', HEAD2: '最小', COL: 'elapsed_time_seconds_min',
         NUMBER_FORMAT: '#,##0.00'},
        {HEAD1: '', HEAD2: '最大', COL: 'elapsed_time_seconds_max',
         NUMBER_FORMAT: '#,##0.00',
         WARNING_VALUE_OVER: 600, WARNING_BACKGROUND_COLOR: 'FF8C00'},  # 処理時間が１０分を超える場合ワーニング
        {HEAD1: '', HEAD2: '合計', COL: 'elapsed_time_seconds',
         NUMBER_FORMAT: '#,##0.00'},
        {HEAD1: '', HEAD2: '平均', COL: 'elapsed_time_seconds_mean',
         NUMBER_FORMAT: '#,##0.00'},
        {HEAD1: 'メモリ使用量(kb)', HEAD2: '最小', COL: 'memusage/max_min',
         DIGIT_ADJUSTMENT: 1000},
        {HEAD1: '', HEAD2: '最大', COL: 'memusage/max_max',
         DIGIT_ADJUSTMENT: 1000,
         WARNING_VALUE_OVER: 300000000, WARNING_BACKGROUND_COLOR: 'FF8C00'},     # メモリ使用量の300mbを超えている場合ワーニング
        {HEAD1: '', HEAD2: '平均', COL: 'memusage/max_mean',
         DIGIT_ADJUSTMENT: 1000},
        {HEAD1: '総リクエスト数', HEAD2: '最小', COL: 'downloader/request_count_min'},
        {HEAD1: '', HEAD2: '最大', COL: 'downloader/request_count_max'},
        {HEAD1: '', HEAD2: '合計', COL: 'downloader/request_count'},
        {HEAD1: '', HEAD2: '平均', COL: 'downloader/request_count_mean',
         NUMBER_FORMAT: '#,##0.00'},
        {HEAD1: '総レスポンス数', HEAD2: '最小', COL: 'downloader/response_count_min'},
        {HEAD1: '', HEAD2: '最大', COL: 'downloader/response_count_max'},
        {HEAD1: '', HEAD2: '合計', COL: 'downloader/response_count'},
        {HEAD1: '', HEAD2: '平均', COL: 'downloader/response_count_mean',
         NUMBER_FORMAT: '#,##0.00'},
        {HEAD1: 'リクエストの', HEAD2: '深さ最大', COL: 'request_depth_max_max'},
        {HEAD1: 'レスポンス量(kb)', HEAD2: '合計', COL: 'downloader/response_bytes',
         DIGIT_ADJUSTMENT: 1000},
        {HEAD1: '', HEAD2: '平均', COL: 'downloader/response_bytes_mean',
         DIGIT_ADJUSTMENT: 1000},
        {HEAD1: 'リトライ件数', HEAD2: '合計', COL: 'retry/count'},
        {HEAD1: '', HEAD2: '平均', COL: 'retry/count_mean',
         NUMBER_FORMAT: '#,##0.00'},
        {HEAD1: '保存件数', HEAD2: '合計', COL: 'item_scraped_count',
         WARNING_VALUE: [0], WARNING_BACKGROUND_COLOR:'FF8C00'},    # 保存件数がゼロの場合ワーニング
        {HEAD1: '', HEAD2: '平均', COL: 'item_scraped_count_mean',
         NUMBER_FORMAT: '#,##0.00'},
    ]

    def run(self, **kwargs):
        ''' '''
        self.run_init()

        self.logger.info(
            f'=== StatsAnalysisReportTask run kwargs : {kwargs}')

        # 入力パラメータのバリデーション
        try:
            self.stats_analysis_report_input = StatsAnalysisReportInput(
                start_time=self.start_time,
                report_term=kwargs[self.stats_analysis_report_input.REPORT_TERM],
                totalling_term=kwargs[self.stats_analysis_report_input.TOTALLING_TERM],
                base_date=kwargs[self.stats_analysis_report_input.BASE_DATE],
                # report_term=kwargs['report_term'],
                # totalling_term=kwargs['totalling_term'],
                # base_date=kwargs['base_date'],
            )
        except ValidationError as e:
            self.logger.error(
                f'=== StatsAnalysisReportTask run : {e.errors()}')
            raise ENDRUN(state=state.Failed())

        self.logger.info(
            f'=== StatsAnalysisReportTask run 基準日from ~ to : {self.stats_analysis_report_input.base_date_get()}')

        # 入力パラメータのレポート期間内の統計情報集計レコードを取得
        stats_info_collect_records, record_count = self.stats_info_collect_get()

        # レコードがあれば
        if record_count:
            # 統計情報集計レコードよりデータフレームを復元する。
            collect_data = StatsInfoCollectData()
            for stats_info_collect_record in stats_info_collect_records:
                collect_data.dataframe_recovery(
                    stats_info_collect_record)

            # 入力パラメータの集計期間単位ごとにpandasによる解析を実行
            spider_result_all_df = collect_data.stats_analysis_exec(
                self.stats_analysis_report_input.datetime_term_list())

            self.logger.info(
                f'=== StatsAnalysisReportTask run  : 統計情報集計レコード解析完了')

            # 統計情報解析結果報告用のワークブックの新規作成
            workbook = Workbook()
            ws: Any = workbook.active  # アクティブなワークシートを選択

            # スパイダー統計解析レポートを編集
            self.stats_analysis_report_edit_header(ws)
            stats_warning_flg: bool = self.stats_analysis_report_edit_body(
                ws, collect_data.spider_list, spider_result_all_df, self.stats_analysis_columns_info,)

            # robotsレスポンスステータス解析レポートを編集
            # シートを追加し選択
            _ = 'robots_response_status'
            workbook.create_sheet(title=_)
            any: Any = workbook[_]
            ws = any
            self.collect_result_analysis_report_edit_header(
                ws, 'robots Analysis Report', self.robots_analysis_columns_info,)
            robots_warning_flg: bool = self.collect_result_analysis_report_edit_body(
                ws, collect_data.spider_list, collect_data.robots_result_df['sum'],
                'robots_response_status', self.robots_analysis_columns_info,)

            # downloaderレスポンスステータス解析レポートを編集
            # シートを追加し選択
            _ = 'downloader_response_status'
            workbook.create_sheet(title=_)
            any: Any = workbook[_]
            ws = any
            # downloaderレスポンスステータス統計解析レポートを編集
            self.collect_result_analysis_report_edit_header(
                ws, 'downloader Analysis Report', self.downloader_analysis_columns_info,)
            downloder_warning_flg: bool = self.collect_result_analysis_report_edit_body(
                ws, collect_data.spider_list, collect_data.downloader_result_df['sum'],
                'downloader_response_status', self.downloader_analysis_columns_info,)

            # レポートファイルを保存
            file_name: str = 'stats_analysis_report.xlsx'
            file_path = os.path.join(DATA_DIR, file_name)
            workbook.save(file_path)

            self.logger.info(
                f'=== StatsAnalysisReportTask run  : レポート対象ファイル保存完了 : {file_path}')

            # メールにレポートファイルを添付して送信
            self.report_mail_create_send(
                stats_warning_flg, robots_warning_flg, downloder_warning_flg, file_path)
        else:
            self.logger.warning(
                f'=== StatsAnalysisReportTask run  : レポート対象のレコードがないため処理をスキップします。')

        # 終了処理
        self.closed()

    def stats_info_collect_get(self) -> tuple[Cursor,int]:
        '''
        mongoDBより指定期間内の統計情報集計レコードを取得して返す。
        '''
        stats_info_collect_model = StatsInfoCollectModel(self.mongo)
        base_date_from, base_date_to = self.stats_analysis_report_input.base_date_get()

        conditions: list = []
        conditions.append({'start_time': {'$gte': base_date_from}})
        conditions.append({'start_time': {'$lt': base_date_to}})
        log_filter: Any = {'$and': conditions}

        record_count:int = stats_info_collect_model.count(filter=log_filter)

        stats_info_collect_records: Cursor = stats_info_collect_model.find(
            filter=log_filter,
            projection={'_id': 0, 'parameter': 0}   # 不要項目を除外
        )

        self.logger.info(
            f'=== 統計情報レポート件数({record_count})')

        return stats_info_collect_records, record_count

    def collect_result_analysis_report_edit_header(
            self, ws: Worksheet,            # レポートの出力先のワークシート
            sheet_name: str,                # レポートの出力先のワークシートのシート名
            analysis_columns_info: list,):  # レポート出力対象カラムに関する情報リスト
        '''集計結果解析レポート用Excelの編集(見出し)'''

        ws.title = sheet_name  # シート名変更

        # 罫線定義
        side = Side(style='thin', color='000000')
        border = Border(top=side, bottom=side, left=side, right=side)
        # 背景色
        fill1 = PatternFill(patternType='solid', fgColor='0066CC')

        for i, col_info in enumerate(analysis_columns_info):
            # 見出し１行目
            any: Any = ws[f'{get_column_letter(i + 1)}{str(1)}']
            head1_cell: Cell = any   # 型ヒントエラー回避用にAnyに一度入れて、改めてCellの項目に渡す
            ws[head1_cell.coordinate] = col_info[self.HEAD1]
            head1_cell.fill = fill1
            head1_cell.border = border
            head1_cell.alignment = Alignment(
                horizontal="center")  # 選択範囲内中央寄せ

    def collect_result_analysis_report_edit_body(
            self,
            ws: Worksheet,                  # レポートの出力先のワークシート
            spider_list: pd.Series,         # スパイダーの一覧
            result_df: pd.DataFrame,        # mongoDBへの集計結果より作成したデータフレーム
            check_col: str,                 # チェックを行いたいカラム
            analysis_columns_info: list,    # レポート出力対象カラムに関する情報リスト
    ) -> bool:
        '''
        集計結果解析レポート用Excelの編集。
        ワーニング有無フラグを返す。(True:有り、False:無し)
        '''

        # 罫線定義
        side = Side(style='thin', color='000000')
        border = Border(top=side, bottom=side, left=side, right=side)

        # 集計期間の数を確認
        aggregate_base_term_list_count = len(
            result_df['aggregate_base_term'].drop_duplicates())

        # ワーニング有無フラグ(初期値：無し)
        warning_flg: bool = False

        # エクセルシートの編集開始行数初期化
        base_row_idx: int = 2
        # スパイダー別に解析を行う。
        for spider in spider_list:
            columns_info_by_spider = deepcopy(analysis_columns_info)
            # スパイダー別のデータフレームを作成
            by_spider_df = result_df.query(
                f'spider_name == "{spider}"')
            # データ無しの件数を確認 ※スパイダーの可動前はデータ無しとなる。
            data_none_count: int = len(
                by_spider_df.query(f'{check_col} == ""'))
            nomal_count: int = aggregate_base_term_list_count - data_none_count

            # ステータス別の一覧を作成し、
            # ステータス別の出現件数が集計期間数と一致するか確認
            status_list_sr: pd.Series = (
                by_spider_df[check_col].drop_duplicates())
            for status in status_list_sr:
                # データ無しは除く
                if status:
                    by_status_df = by_spider_df.query(
                        f'{check_col} == "{status}"')
                    if nomal_count != len(by_status_df):
                        for col_info in columns_info_by_spider:
                            if col_info[self.COL] == check_col:
                                col_info[self.WARNING_VALUE].append(status)

            # 列ごとにエクセルに編集
            for col_idx, col_info in enumerate(columns_info_by_spider):
                # データフレームの列ごとに１行ずつ処理を実施
                for row_idx, value in enumerate(by_spider_df[col_info[self.COL]]):
                    # 更新対象のセルに値を設定
                    f'{get_column_letter(col_idx + 1)}{str(base_row_idx + row_idx)}'
                    # any: Any = ws[get_column_letter(col_idx + 1) + str(base_row_idx + row_idx)]
                    any: Any = ws[f'{get_column_letter(col_idx + 1)}{str(base_row_idx + row_idx)}']
                    target_cell: Cell = any
                    ws[target_cell.coordinate] = value

                    # 同値カラー調整列の場合、更新対象セルの上と同値ならば色を変える。
                    if self.EQUIVALENT_COLOR in col_info:
                        any: Any = ws[f'{get_column_letter(col_idx + 1)}{str(base_row_idx + row_idx - 1)}']
                        compare_cell: Cell = any
                        # compare_cell: Cell = ws[get_column_letter(
                        #     col_idx + 1) + str(base_row_idx + row_idx - 1)]
                        if target_cell.value == compare_cell.value:
                            target_cell.font = Font(
                                color=col_info[self.EQUIVALENT_COLOR])
                            warning_flg = True

                    # ワーニング値チェック列の場合、ワーニング値セルの背景色を設定。
                    if self.WARNING_VALUE in col_info:
                        if value in col_info[self.WARNING_VALUE]:
                            target_cell.fill = PatternFill(
                                patternType='solid', fgColor=col_info[self.WARNING_BACKGROUND_COLOR])
                            warning_flg = True

            # １つのスパイダーに対する編集が終わった段階で、次のスパイダー用にエクセルの編集行の基準を設定する。
            base_row_idx = base_row_idx + len(by_spider_df)

        # 各明細行のセルに罫線を設定する。
        max_cell: str = get_column_letter(
            ws.max_column) + str(ws.max_row)  # "BC55"のようなセル番地を生成
        # cell: Cell  # 型ヒントの定義のみ
        # for cells in ws[f'a2:{max_cell}']:
        #     for cell in cells:
        #         cell.border = border
        cells = ws[f'a2:{max_cell}']
        if type(cells) is tuple:
            for cell in cells:
                cell.border = border



        # 列ごとに次の処理を行う。
        # 最大幅を確認
        # それに合わせた幅を設定する。
        # for cols in ws.columns:
        for cols in ws.iter_cols():
            max_length = 0
            column = cols[0].column_letter  # 列名A,Bなどを取得
            # column = cols.column_letter  # 列名A,Bなどを取得
            for cell in cols:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            # #型ヒントでcolumn_dimensionsが存在しないものとみなされエラーが出るため、動的メソッドの実行形式で記述
            # ws.column_dimensions[column].width = (max_length + 2.07)
            getattr(ws, 'column_dimensions')()[column].width = (max_length + 2.07)

        # ウィンドウ枠の固定。１行２列は常に表示させる。
        ws.freeze_panes = 'c2'

        return warning_flg

    def stats_analysis_report_edit_header(self, ws: Worksheet):
        '''統計情報解析レポート用Excelの見出し編集'''
        ws.title = 'Stats Analysis Report'  # シート名変更

        # 罫線定義
        side = Side(style='thin', color='000000')
        border = Border(top=side, bottom=side, left=side, right=side)
        fill1 = PatternFill(patternType='solid', fgColor='0066CC')
        fill2 = PatternFill(patternType='solid', fgColor='0099CC')

        for i, col_info in enumerate(self.stats_analysis_columns_info):
            # 見出し１行目
            any: Any = ws[f'{get_column_letter(i + 1)}{str(1)}']
            head1_cell: Cell = any
            ws[head1_cell.coordinate] = col_info[self.HEAD1]
            head1_cell.fill = fill1
            head1_cell.border = border
            head1_cell.alignment = Alignment(
                horizontal="centerContinuous")  # 選択範囲内中央寄せ

            # 見出し２行目
            any: Any = ws[f'{get_column_letter(i + 1)}{str(2)}']
            head2_cell: Cell = any
            ws[head2_cell.coordinate] = col_info[self.HEAD2]
            head2_cell.fill = fill2
            head2_cell.border = border
            head2_cell.alignment = Alignment(horizontal="center")  # 中央寄せ

    def stats_analysis_report_edit_body(
        self,
        ws: Worksheet,
        spider_list: pd.Series,         # スパイダーの一覧
        result_df: pd.DataFrame,
        analysis_columns_info: list,
    ) -> bool:
        '''
        統計情報解析レポート用Excelの編集。
        ワーニング有無フラグを返す。(True:有り、False:無し)
        '''
        # 罫線定義
        side = Side(style='thin', color='000000')
        border = Border(top=side, bottom=side, left=side, right=side)

        # 集計期間の数を確認
        aggregate_base_term_list_count = len(
            result_df['aggregate_base_term'].drop_duplicates())

        # ワーニング有無フラグ(初期値：無し)
        warning_flg: bool = False

        # エクセルシートの編集開始行数初期化
        base_row_idx: int = 3
        # スパイダー別に解析を行う。
        for spider in spider_list:
            columns_info_by_spider = deepcopy(analysis_columns_info)
            # スパイダー別のデータフレームを作成
            by_spider_df = result_df.query(f'spider_name == "{spider}"')
            # データ無しの件数を確認 ※スパイダーの可動前はデータ無しとなる。
            data_none_count: int = len(
                by_spider_df.query(f'elapsed_time_seconds == ""'))
            nomal_count: int = aggregate_base_term_list_count - data_none_count

            # 列ごとにエクセルに編集
            for col_idx, col_info in enumerate(columns_info_by_spider):
                for row_idx, value in enumerate(by_spider_df[col_info[self.COL]]):
                    # 更新対象のセル
                    any: Any = ws[f'{get_column_letter(col_idx + 1)}{str(base_row_idx + row_idx)}']
                    target_cell: Cell = any

                    # 表示単位の切り上げ (example:b -> kb)
                    custom_value = value
                    if self.DIGIT_ADJUSTMENT in col_info:
                        if value:
                            custom_value = int(
                                value) // col_info[self.DIGIT_ADJUSTMENT]
                            Decimal(str(custom_value)).quantize(
                                Decimal('0'), rounding=ROUND_HALF_UP)

                    # 小数点以下の調整
                    if self.NUMBER_FORMAT in col_info:
                        target_cell.number_format = col_info[self.NUMBER_FORMAT]
                    else:
                        target_cell.number_format = '#,##0'

                    # 更新対象のセルに値を設定
                    ws[target_cell.coordinate] = custom_value

                    # 同値カラー調整
                    if self.EQUIVALENT_COLOR in col_info:
                        # 比較用の１つ上のセルと同じ値の場合は文字色を変更
                        any: Any = ws[f'{get_column_letter(col_idx + 1)}{str(base_row_idx + row_idx - 1)}']
                        compare_cell: Cell = any
                        if target_cell.value == compare_cell.value:
                            target_cell.font = Font(
                                color=col_info[self.EQUIVALENT_COLOR])

                    # ワーニング値チェック列の場合、ワーニング値セルの背景色を設定。
                    if self.WARNING_VALUE in col_info:
                        if value in col_info[self.WARNING_VALUE]:
                            target_cell.fill = PatternFill(
                                patternType='solid', fgColor=col_info[self.WARNING_BACKGROUND_COLOR])
                            warning_flg = True

                    # ワーニング値超過チェック列の場合、ワーニング値超過セルの背景色を設定。
                    if self.WARNING_VALUE_OVER in col_info:
                        if value:
                            if int(value) > col_info[self.WARNING_VALUE_OVER]:
                                target_cell.fill = PatternFill(
                                    patternType='solid', fgColor=col_info[self.WARNING_BACKGROUND_COLOR])
                                warning_flg = True

            # １つのスパイダーに対する編集が終わった段階で、次のスパイダー用にエクセルの編集行の基準を設定する。
            base_row_idx = base_row_idx + len(by_spider_df)

        # 各明細行のセルに罫線を設定する。
        max_cell: str = get_column_letter(
            ws.max_column) + str(ws.max_row)  # "BC55"のようなセル番地を生成
        cells = ws[f'a3:{max_cell}']
        if type(cells) is tuple:
            for cell in cells:
                cell: Cell
                cell.border = border

        # 列ごとに次の処理を行う。
        # 最大幅を確認
        # それに合わせた幅を設定する。
        # for col in ws.columns:
        for col in ws.iter_cols():
            max_length = 0
            column = col[0].column_letter  # 列名A,Bなどを取得
            for cell in col:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            # 型ヒントでcolumn_dimensionsが存在しないものとみなされエラーが出るため、動的メソッドの実行形式で記述
            # ws.column_dimensions[column].width = (max_length + 2.07)
            getattr(ws, 'column_dimensions')()[column].width = (max_length + 2.07)

        # ウィンドウ枠の固定。２行２列は常に表示させる。
        ws.freeze_panes = 'c3'

        return warning_flg

    def report_mail_create_send(self, stats_warning_flg: bool, robots_warning_flg: bool, downloder_warning_flg: bool, file_path: str):
        '''メールにレポートファイルを添付して送信'''
        base_date_from, base_date_to = self.stats_analysis_report_input.base_date_get()

        title = 'stats_analysis_report'
        if stats_warning_flg or robots_warning_flg or downloder_warning_flg:
            title = title + '(ワーニング有り)'

        warning_messege: str = ''
        if stats_warning_flg:
            warning_messege = '<p>statsでワーニング発生</p>'
        if robots_warning_flg:
            warning_messege = f'{warning_messege}<p>robots response statusでワーニング発生</p>'
        if downloder_warning_flg:
            warning_messege = f'{warning_messege}<p>downloder response statusでワーニング発生</p>'

        messege = f'''
        <html>
            <body>
                <p>各種実行結果を解析したレポート</p>
                <p>=== 実行条件 ============================================================</p>
                <p>start_time = {self.start_time.isoformat()}</p>
                <p>base_date_from = {base_date_from.isoformat()}</p>
                <p>base_date_to = {base_date_to.isoformat()}</p>
                <p>report_term = {self.stats_analysis_report_input.report_term}</p>
                <p>totalling_term = {self.stats_analysis_report_input.totalling_term}</p>
                <p>=========================================================================</p>
                {warning_messege}
            </body>
        </html>'''
        # メール送信
        mail_attach_send(title=title, msg=messege, filepath=file_path)
