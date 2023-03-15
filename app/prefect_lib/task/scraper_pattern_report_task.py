import os
import sys
from typing import Any, Final
import pandas as pd
from pydantic import ValidationError
from prefect.engine import state
from prefect.engine.runner import ENDRUN
from pymongo.cursor import Cursor
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
from openpyxl.chart.bar_chart import BarChart
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter
path = os.getcwd()
sys.path.append(path)
from prefect_lib.data_models.scraper_pattern_report_data import ScraperPatternReportData
from prefect_lib.data_models.scraper_pattern_report_input import ScraperPatternReportInput
from prefect_lib.task.extentions_task import ExtensionsTask
from BrownieAtelierMongo.data_models.scraper_info_by_domain_data import ScraperInfoByDomainData
from BrownieAtelierMongo.collection_models.scraper_info_by_domain_model import ScraperInfoByDomainModel
from BrownieAtelierMongo.collection_models.news_clip_master_model import NewsClipMasterModel
from BrownieAtelierNotice.mail_attach_send import mail_attach_send
from shared.settings import DATA_DIR


class ScrapyingPatternReportTask(ExtensionsTask):
    '''スクレイパーパターン情報の使用状況のレポートを作成する'''
    # 作成するレポートの見出し項目
    HEAD1: Final[str] = 'head1'
    HEAD2: Final[str] = 'head2'
    COL: Final[str] = 'col'
    DOMAIN: Final[str] = 'domain'
    EQUIVALENT_COLOR: Final[str] = 'equivalent_color'
    SCRAPER_ITEM: Final[str] = 'scraper_item'
    PRIORITY: Final[str] = 'priority'
    PATTERN: Final[str] = 'pattern'

    scraper_pattern_analysis_columns_info: list = [
        # head1                     : 必須 : 見出し１行目
        # head2                     : 必須 : 見出し２行目
        # col                       : 必須 : データフレーム列名
        # digit_adjustment          : 任意 : 単位の調整。1000とした場合、'value/1000'となる。
        # number_format             : 任意 : 小数点以下の桁数。省略した場合'#,##0'
        # number_format             : 任意 : 小数点以下の桁数。省略した場合'#,##0'
        # equivalent_color          : 任意 : 上のセルと同値の場合の文字色。上と同値の場合、文字色を薄くするなどに使用する。
        # warning_value             : 任意 : ワーニングとする値を入れる配列
        # warning_value_over        : 任意 : 超過したらワーニングとする値
        # warning_background_color  : 任意 : ワーンングとなったセルの背景色
        #{'head1': '集計期間', 'head2': '', 'col': 'aggregate_base_term'},
        {HEAD1: 'スクレイパー情報', HEAD2: 'ドメイン', COL: DOMAIN,
         EQUIVALENT_COLOR: 'bbbbbb'},
        {HEAD1: '', HEAD2: 'アイテム', COL: SCRAPER_ITEM,
         EQUIVALENT_COLOR: 'bbbbbb'},
        {HEAD1: '', HEAD2: '優先順位', COL: PRIORITY},
        {HEAD1: '', HEAD2: 'パターン', COL: PATTERN},
    ]

    def run(self, **kwargs):
        ''' '''
        self.run_init()

        self.logger.info(
            f'=== ScrapyingPatternReportTask run kwargs : {kwargs}')

        # 入力パラメータのバリデーション
        try:
            self.scraper_pattern_report_input = ScraperPatternReportInput(
                start_time=self.start_time,
                report_term=kwargs[self.scraper_pattern_report_input.REPORT_TERM],
                base_date=kwargs[self.scraper_pattern_report_input.BASE_DATE],
            )
        except ValidationError as e:
            self.logger.error(
                f'=== ScrapyingPatternReportTask run : {e.errors()}')
            raise ENDRUN(state=state.Failed())

        self.logger.info(
            f'=== ScrapyingPatternReportTask run 基準日from ~ to : {self.scraper_pattern_report_input.base_date_get()}')

        news_clip_master = NewsClipMasterModel(self.mongo)
        scraper_pattern_report_data = ScraperPatternReportData()

        # ドメイン別スクレイパー情報を順に処理する。
        base_date_from, base_date_to = self.scraper_pattern_report_input.base_date_get()
        for scraper_info_by_domain_data in self.scraper_info_by_domain_get():

            # 対象件数を確認
            conditions: list = []
            conditions.append(
                {NewsClipMasterModel.CRAWLING_START_TIME: {'$gte': base_date_from}})
            conditions.append({NewsClipMasterModel.CRAWLING_START_TIME: {'$lt': base_date_to}})
            conditions.append(
                {NewsClipMasterModel.DOMAIN: scraper_info_by_domain_data.domain_get()})
            filter: dict = {'$and': conditions}
            record_count = news_clip_master.count(filter=filter)
            self.logger.info(
                f'=== ScrapyingPatternReportTask run news_clip_master domain = {scraper_info_by_domain_data.domain_get()},  件数 = {str(record_count)}')

            # データフレーム（マスター）を作成
            skeleton_for_counters: list = scraper_info_by_domain_data.making_into_a_table_format()
            for skeleton_for_counter in skeleton_for_counters:
                scraper_pattern_report_data.scraper_info_master_store(
                    skeleton_for_counter)

            # news_clip_masterからレコードを順に取得する。
            # 取得したレコードをデータフレーム（カウント用）へ挿入
            for master_record in news_clip_master.limited_find(
                    filter=filter, projection={
                        NewsClipMasterModel.DOMAIN: 1,
                        NewsClipMasterModel.CRAWLING_START_TIME: 1,
                        NewsClipMasterModel.PATTERN: 1}):

                pattern_info: dict = master_record[NewsClipMasterModel.DOMAIN]

                for pattern_key, pattern_value in pattern_info.items():
                    scraper_pattern_report_data.scraper_info_counter_store({
                        scraper_pattern_report_data.DOMAIN: master_record[NewsClipMasterModel.DOMAIN],
                        scraper_pattern_report_data.SCRAPER_ITEM: pattern_key,
                        scraper_pattern_report_data.PATTERN: pattern_value,
                        scraper_pattern_report_data.COUNT_OF_USE: 1,
                    })

        # 集計対象のnews_clip_masterが1件もない場合処理を中止する。
        if len(scraper_pattern_report_data.scraper_pattern_counter_df) == 0:
            self.logger.warning(
                f'=== ScrapyingPatternReportTask run  : レポート対象のレコードがないため処理をスキップします。')
            raise ENDRUN(state=state.Failed())

        # データフレーム（カウント用）とデータフレーム（マスター）を使用し、
        # データフレーム（結果）を生成する。
        scraper_pattern_report_data.scraper_info_analysis()

        self.logger.info(
            f'=== ScrapyingPatternReportTask run  : スクレイパー情報集計完了')

        # スクレイパー情報集計結果報告用のワークブックの新規作成
        workbook = Workbook()
        ws: Any = workbook.active  # アクティブなワークシートを選択

        # スクレイパー情報集計結果報告用レポートを編集
        self.scraper_pattern_analysis_report_edit_header(ws)
        self.scraper_pattern_analysis_report_edit_body(
            ws, scraper_pattern_report_data.result_df)

        # レポートファイルを保存
        file_name: str = 'scraper_pattern_analysis_report.xlsx'
        file_path = os.path.join(DATA_DIR, file_name)
        workbook.save(file_path)

        self.logger.info(
            f'=== ScrapyingPatternReportTask run  : レポート対象ファイル保存完了 : {file_path}')

        # メールにレポートファイルを添付して送信
        self.report_mail_create_send(file_path)

        # 終了処理
        self.closed()

    def scraper_info_by_domain_get(self) -> list[ScraperInfoByDomainData]:
        '''
        mongoDBよりドメイン別スクレイパー情報を全件取得
        '''
        scraper_info_by_domain = ScraperInfoByDomainModel(self.mongo)

        return scraper_info_by_domain.find_and_data_models_get()

    def scraper_pattern_analysis_report_edit_header(self, ws: Worksheet):
        '''スクレイパー情報解析レポート用Excelの見出し編集'''
        ws.title = 'Scraper Info By Domain'  # シート名変更

        # 罫線定義
        side = Side(style='thin', color='000000')
        border = Border(top=side, bottom=side, left=side, right=side)
        fill1 = PatternFill(patternType='solid', fgColor='0066CC')
        fill2 = PatternFill(patternType='solid', fgColor='0099CC')

        for i, col_info in enumerate(self.scraper_pattern_analysis_columns_info):
            # 見出し１行目
            any: Any = ws[f'{get_column_letter(i + 1)}{str(1)}']
            head1_cell: Cell = any
            ws[head1_cell.coordinate] = col_info[self.HEAD1]
            head1_cell.fill = fill1
            head1_cell.border = border
            head1_cell.alignment = Alignment(
                horizontal="centerContinuous")  # 選択範囲内中央寄せ

            # 見出し２行目
            any: Any = ws[f'{get_column_letter(i + 1)}{str(2)}']
            head2_cell: Cell = any
            ws[head2_cell.coordinate] = col_info[self.HEAD2]
            head2_cell.fill = fill2
            head2_cell.border = border
            head2_cell.alignment = Alignment(horizontal="center")  # 中央寄せ

    def scraper_pattern_analysis_report_edit_body(self, ws: Worksheet, result_df: pd.DataFrame):
        '''
        スクレイパー情報解析レポート用Excelの編集。
        '''
        # 罫線定義
        side = Side(style='thin', color='000000')
        border = Border(top=side, bottom=side, left=side, right=side)

        # エクセルシートの編集開始行数初期化
        base_row_idx: int = 3
        # 列ごとにエクセルに編集
        for col_idx, col_info in enumerate(self.scraper_pattern_analysis_columns_info):
            for row_idx, value in enumerate(result_df[col_info[self.COL]]):
                # 更新対象のセル
                any: Any = ws[f'{get_column_letter(col_idx + 1)}{str(base_row_idx + row_idx)}']
                target_cell: Cell = any

                # 更新対象のセルに値を設定
                ws[target_cell.coordinate] = value

                # 同値カラー調整
                if self.EQUIVALENT_COLOR in col_info:
                    # 比較用の１つ上のセルと同じ値の場合は文字色を変更
                    any: Any = ws[f'{get_column_letter(col_idx + 1)}{str(base_row_idx + row_idx - 1)}']
                    compare_cell: Cell = any
                    if target_cell.value == compare_cell.value:
                        target_cell.font = Font(
                            color=col_info[self.EQUIVALENT_COLOR])

        # 各明細行のセルに罫線を設定する。
        max_cell: str = get_column_letter(
            ws.max_column) + str(ws.max_row)  # "BC55"のようなセル番地を生成
        cells = ws[f'a3:{max_cell}']
        if type(cells) is tuple:
            for cell in cells:
                cell.border = border

        # 列ごとに次の処理を行う。
        # 最大幅を確認
        # それに合わせた幅を設定する。
        for col in ws.iter_cols():
            max_length = 0
            column = col[0].column_letter  # 列名A,Bなどを取得
            for cell in col:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))

            # 型ヒントでcolumn_dimensionsが存在しないものとみなされエラーが出るため、動的メソッドの実行形式で記述
            getattr(ws, 'column_dimensions')()[column].width = (max_length + 2.2)

        # ウィンドウ枠の固定。２行２列は常に表示させる。
        ws.freeze_panes = 'c3'

    def report_mail_create_send(self, file_path: str):
        '''メールにレポートファイルを添付して送信'''
        base_date_from, base_date_to = self.scraper_pattern_report_input.base_date_get()

        title = 'scraper_pattern_analysis_report'

        messege = f'''
        <html>
            <body>
                <p>各種実行結果を解析したレポート</p>
                <p>=== 実行条件 ============================================================</p>
                <p>start_time = {self.start_time.isoformat()}</p>
                <p>base_date_from = {base_date_from.isoformat()}</p>
                <p>base_date_to = {base_date_to.isoformat()}</p>
                <p>report_term = {self.scraper_pattern_report_input.report_term}</p>
                <p>=========================================================================</p>
            </body>
        </html>'''
        # メール送信
        mail_attach_send(title=title, msg=messege, filepath=file_path)
