from __future__ import annotations
import os
import sys
from typing import Any, Sequence
from datetime import datetime
from pydantic import ValidationError
from prefect.engine import state
from prefect.engine.runner import ENDRUN
from pymongo.cursor import Cursor
from urllib.parse import urlparse

from openpyxl import Workbook
from openpyxl.chart.bar_chart import BarChart
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

path = os.getcwd()
sys.path.append(path)
from prefect_lib.task.extentions_task import ExtensionsTask
from common_lib.timezone_recovery import timezone_recovery
from common_lib.mail_send import mail_send
from prefect_lib.data_models.log_analysis_report_model import LogAnalysisReportModel
from models.mongo_model import MongoModel
from models.asynchronous_report_model import AsynchronousReportModel
from models.crawler_logs_model import CrawlerLogsModel


class LogAnalysisReportTask(ExtensionsTask):
    '''
    '''

    def run(self, **kwargs):
        '''ここがprefectで起動するメイン処理'''

        self.start_time
        self.mongo
        self.logger.info(
            f'=== LogAnalysisReportTask run kwargs : {kwargs}')

        try:
            log_analysis_report = LogAnalysisReportModel(
                start_time=self.start_time,
                report_term=kwargs['report_term'],
                base_date=kwargs['base_date'],
            )
        except ValidationError as e:
            # e.json()エラー結果をjson形式で見れる。
            # e.errors()エラー結果をdict形式で見れる。
            # str(e)エラー結果をlist形式で見れる。
            self.logger.error(
                f'=== LogAnalysisReportTask run : {e.errors()}')
            raise ENDRUN(state=state.Failed())

        self.logger.info(
                f'=== LogAnalysisReportTask run 基準日from ~ to : {log_analysis_report.base_date_get()}')

        '''
        asynchronous_report
        crawler_logs
        '''
        self.asynchronous_report_analysis(log_analysis_report)
        self.crawler_logs_analysis(log_analysis_report)

        # 各コレクション、solrの件数を取得して同期確認
        # ここで上記の解析結果よりレポートを作成する？

        # 終了処理
        self.closed()
        # return ''

    def asynchronous_report_analysis(self, log_analysis_report: LogAnalysisReportModel):
        '''
        まずデータの有無。
        指定期間内に非同期データがあればレポート要。
        record_type、start_time、async_listの3項目。
        record_typeは3種 : news_crawl_async, news_clip_master_async, solr_news_clip_async。
        async_listから総件数、ドメイン別の件数。
        '''

        asynchronous_report_model = AsynchronousReportModel(self.mongo)

        base_date_from, base_date_to = log_analysis_report.base_date_get()

        #
        conditions: list = []
        conditions.append(
            {'start_time': {'$gte': base_date_from}})
        conditions.append(
            {'start_time': {'$lt': base_date_to}})

        log_filter: Any = {'$and': conditions}

        asynchronous_report_records: Cursor = asynchronous_report_model.find(
            filter=log_filter,
            projection={'_id': 0, 'parameter': 0}
        )
        self.logger.info(
            f'=== 非同期レポート件数({asynchronous_report_records.count()})')

        # レコードタイプ別カウントエリア
        by_record_type_count: dict[str, int] = {
            'news_crawl_async': 0,
            'news_clip_master_async': 0,
            'solr_news_clip_async': 0,
        }
        # レコードタイプ別・ドメイン別カウントエリア
        by_record_type_by_domain_count: dict[str, dict[str, int]] = {
            'news_crawl_async': {},
            'news_clip_master_async': {},
            'solr_news_clip_async': {},
        }

        for asynchronous_report_record in asynchronous_report_records:
            # print(asynchronous_report_record)

            # レコードタイプ別に集計を行う。
            by_record_type_count[asynchronous_report_record['record_type']] += 1
            self.by_domain_asynchronous_report_count(
                asynchronous_report_record['async_list'], by_record_type_by_domain_count[asynchronous_report_record['record_type']])

        print(by_record_type_count)
        print(by_record_type_by_domain_count)

        wb = Workbook()
        ws = wb.active
        ws['a1'] = 10
        cell = ws['a1']
        #cell.font = Font()

        wb.save('test.xlsx')

    def by_domain_asynchronous_report_count(self, async_list: list, by_domain_count: dict):
        for url in async_list:
            url_parse = urlparse(url)
            if url_parse.netloc not in by_domain_count:
                by_domain_count[url_parse.netloc] = 0
            by_domain_count[url_parse.netloc] += 1

    def crawler_logs_analysis(self, log_analysis_report: LogAnalysisReportModel):
        '''
        ログレベルワーニング、エラー、クリティカルの発生件数
        record_type = spider_reports
            start_time  record_type domain  spider_name stats
        record_type = その他（タスク）※実行ログ
        '''
        crawler_logs = CrawlerLogsModel(self.mongo)

        base_date_from, base_date_to = log_analysis_report.base_date_get()

        #
        conditions: list = []
        conditions.append(
            {'start_time': {'$gte': base_date_from}})
        conditions.append(
            {'start_time': {'$lt': base_date_to}})

        log_filter: Any = {'$and': conditions}

        crawler_logs_records: Cursor = crawler_logs.find(
            filter=log_filter,
            projection={'_id': 0, 'crawl_urls_list': 0}  # crawl_urls_listは不要
        )
        self.logger.info(
            f'=== ログ件数({crawler_logs_records.count()})')

        print('=============')
        print('=============')
        print('=============')
        # for crawler_logs_record in crawler_logs_records:
        #    print(crawler_logs_record)
