from __future__ import annotations
from copy import deepcopy
import os
import sys
from typing import Any
import pandas as pd
from pydantic import ValidationError
from prefect.engine import state
from prefect.engine.runner import ENDRUN
from pymongo.cursor import Cursor
from decimal import Decimal, ROUND_HALF_UP
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
from openpyxl.chart.bar_chart import BarChart
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter

path = os.getcwd()
sys.path.append(path)
from prefect_lib.settings import DATA_DIR
from common_lib.timezone_recovery import timezone_recovery
from common_lib.mail_attach_send import mail_attach_send
from models.mongo_model import MongoModel
from models.stats_info_collect_model import StatsInfoCollectModel
from prefect_lib.task.extentions_task import ExtensionsTask
from prefect_lib.data_models.stats_analysis_report_input import StatsAnalysisReportInput
from prefect_lib.data_models.stats_info_collect_data import StatsInfoCollectData


class StatsAnalysisReportTask(ExtensionsTask):
    ''''''
    robots_analysis_columns_info: list = [
        # head1                     : 必須 : 見出し１行目
        # col                       : 必須 : データフレーム列名
        # equivalent_color          : 任意 : 上のセルと同値の場合の文字色。上と同値の場合、文字色を薄くするなどに使用する。
        # warning_value             : 任意 : ワーニングとする値を入れる配列
        # warning_background_color  : 任意 : ワーンングとなったセルの背景色
        {'head1': '集計日〜', 'col': 'aggregate_base_term'},
        {'head1': 'スパイダー名', 'col': 'spider_name',
         'equivalent_color': 'bbbbbb'},
        {'head1': 'status', 'col': 'robots_response_status',
         'warning_value': [], 'warning_background_color':'FF8C00'},
        {'head1': 'count', 'col': 'count'},
    ]

    downloader_analysis_columns_info: list = [
        # head1                     : 必須 : 見出し１行目
        # col                       : 必須 : データフレーム列名
        # equivalent_color          : 任意 : 上のセルと同値の場合の文字色。上と同値の場合、文字色を薄くするなどに使用する。
        # warning_value             : 任意 : ワーニングとする値を入れる配列
        # warning_background_color  : 任意 : ワーンングとなったセルの背景色
        {'head1': '集計日〜', 'col': 'aggregate_base_term'},
        {'head1': 'スパイダー名', 'col': 'spider_name',
         'equivalent_color': 'bbbbbb'},
        {'head1': 'status', 'col': 'downloader_response_status',
         'warning_value': [], 'warning_background_color':'FF8C00'},
        {'head1': 'count', 'col': 'count'},
    ]

    stats_analysis_columns_info: list = [
        # head1             : 必須 : 見出し１行目
        # head2             : 必須 : 見出し２行目
        # col               : 必須 : データフレーム列名
        # digit_adjustment  : 任意 : 単位の調整。1000とした場合、'value/1000'となる。
        # number_format     : 任意 : 小数点以下の桁数。省略した場合'#,##0'
        # number_format     : 任意 : 小数点以下の桁数。省略した場合'#,##0'
        # equivalent_color  : 任意 : 上のセルと同値の場合の文字色。上と同値の場合、文字色を薄くするなどに使用する。
        {'head1': '集計日〜', 'head2': '', 'col': 'aggregate_base_term'},
        {'head1': 'スパイダー名', 'head2': '', 'col': 'spider_name',
         'equivalent_color': 'bbbbbb'},
        {'head1': 'ログレベル件数', 'head2': 'CRITICAL', 'col': 'log_count/CRITICAL'},
        {'head1': '', 'head2': 'ERROR', 'col': 'log_count/ERROR'},
        {'head1': '', 'head2': 'WARNING', 'col': 'log_count/WARNING'},
        {'head1': '処理時間(秒)', 'head2': '最小', 'col': 'elapsed_time_seconds_min',
         'number_format': '#,##0.00'},
        {'head1': '', 'head2': '最大', 'col': 'elapsed_time_seconds_max',
         'number_format': '#,##0.00'},
        {'head1': '', 'head2': '合計', 'col': 'elapsed_time_seconds',
         'number_format': '#,##0.00'},
        {'head1': '', 'head2': '平均', 'col': 'elapsed_time_seconds_mean',
         'number_format': '#,##0.00'},
        {'head1': 'メモリ使用量(kb)', 'head2': '最小', 'col': 'memusage/max_min',
         'digit_adjustment': 1000},
        {'head1': '', 'head2': '最大', 'col': 'memusage/max_max',
         'digit_adjustment': 1000},
        {'head1': '', 'head2': '平均', 'col': 'memusage/max_mean',
         'digit_adjustment': 1000},
        {'head1': '総リクエスト数', 'head2': '最小', 'col': 'downloader/request_count_min'},
        {'head1': '', 'head2': '最大', 'col': 'downloader/request_count_max'},
        {'head1': '', 'head2': '合計', 'col': 'downloader/request_count'},
        {'head1': '', 'head2': '平均', 'col': 'downloader/request_count_mean',
         'number_format': '#,##0.00'},
        {'head1': '総レスポンス数', 'head2': '最小', 'col': 'downloader/response_count_min'},
        {'head1': '', 'head2': '最大', 'col': 'downloader/response_count_max'},
        {'head1': '', 'head2': '合計', 'col': 'downloader/response_count'},
        {'head1': '', 'head2': '平均', 'col': 'downloader/response_count_mean',
         'number_format': '#,##0.00'},
        {'head1': 'リクエストの', 'head2': '深さ最大', 'col': 'request_depth_max_max'},
        {'head1': 'レスポンス量(kb)', 'head2': '合計', 'col': 'downloader/response_bytes',
         'digit_adjustment': 1000},
        {'head1': '', 'head2': '平均', 'col': 'downloader/response_bytes_mean',
         'digit_adjustment': 1000},
        {'head1': 'リトライ件数', 'head2': '合計', 'col': 'retry/count'},
        {'head1': '', 'head2': '平均', 'col': 'retry/count_mean',
         'number_format': '#,##0.00'},
        {'head1': '保存件数', 'head2': '合計', 'col': 'item_scraped_count'},
        {'head1': '', 'head2': '平均', 'col': 'item_scraped_count_mean',
         'number_format': '#,##0.00'},
    ]

    def run(self, **kwargs):
        '''ここがprefectで起動するメイン処理'''

        self.start_time
        self.mongo
        self.logger.info(
            f'=== StatsAnalysisReportTask run kwargs : {kwargs}')

        try:
            self.stats_analysis_report_input = StatsAnalysisReportInput(
                start_time=self.start_time,
                report_term=kwargs['report_term'],
                totalling_term=kwargs['totalling_term'],
                base_date=kwargs['base_date'],
            )
        except ValidationError as e:
            self.logger.error(
                f'=== StatsAnalysisReportTask run : {e.errors()}')
            raise ENDRUN(state=state.Failed())

        self.logger.info(
            f'=== StatsAnalysisReportTask run 基準日from ~ to : {self.stats_analysis_report_input.base_date_get()}')

        stats_info_collect_records: Cursor = self.stats_info_collect_get()

        self.stats_analysis_report_create(stats_info_collect_records)

        # 終了処理
        self.closed()

    def stats_info_collect_get(self) -> Cursor:
        '''
        mongoDBより指定期間内の統計情報集計レコードを取得して返す。
        '''
        stats_info_collect_model = StatsInfoCollectModel(self.mongo)
        base_date_from, base_date_to = self.stats_analysis_report_input.base_date_get()

        conditions: list = []
        conditions.append(
            {'start_time': {'$gte': base_date_from}})
        conditions.append(
            {'start_time': {'$lt': base_date_to}})

        log_filter: Any = {'$and': conditions}

        stats_info_collect_records: Cursor = stats_info_collect_model.find(
            filter=log_filter,
            projection={'_id': 0, 'parameter': 0}
        )
        self.logger.info(
            f'=== 統計情報レポート件数({stats_info_collect_records.count()})')

        return stats_info_collect_records

    def stats_analysis_report_create(self,stats_info_collect_records:Cursor):
        '''  '''
        collect_data = StatsInfoCollectData()
        for stats_info_collect_record in stats_info_collect_records:
            collect_data.dataframe_recovery(
                stats_info_collect_record)

        # 集計期間リストごとに解析を実行
        spider_result_all_df = collect_data.stats_analysis_exec(
            self.stats_analysis_report_input.datetime_term_list())

        # ワークブックの新規作成
        workbook = Workbook()
        # アクティブなワークシートを選択
        ws: Worksheet = workbook.active
        # スパイダー統計解析レポートを編集
        self.stats_analysis_report_edit_header(ws)
        self.stats_analysis_report_edit_body(ws, spider_result_all_df)

        # シートを追加し選択
        _ = 'robots_response_status'
        workbook.create_sheet(title=_)
        any: Any = workbook[_]
        ws = any
        self.collect_result_analysis_report_edit_header(
            ws,
            'robots Analysis Report',
            self.robots_analysis_columns_info,)
        self.collect_result_analysis_report_edit_body(
            ws,
            collect_data.spider_list,
            collect_data.robots_result_df['sum'],
            'robots_response_status',
            self.robots_analysis_columns_info,)

        # シートを追加し選択
        _ = 'downloader_response_status'
        workbook.create_sheet(title=_)
        any: Any = workbook[_]
        ws = any
        # downloaderレスポンスステータス統計解析レポートを編集
        self.collect_result_analysis_report_edit_header(
            ws,
            'downloader Analysis Report',
            self.downloader_analysis_columns_info,)
        self.collect_result_analysis_report_edit_body(
            ws,
            collect_data.spider_list,
            collect_data.downloader_result_df['sum'],
            'downloader_response_status',
            self.downloader_analysis_columns_info,)

        # レポートファイルの保存
        file_name: str = 'stats_analysis_report.xlsx'
        file_path = os.path.join(DATA_DIR, file_name)
        workbook.save(file_path)

        # メールにレポートファイルを添付して送信
        base_date_from, base_date_to = self.stats_analysis_report_input.base_date_get()
        title = 'stats_analysis_report'
        msg = f'''
        <html>
            <body>
                <p>各種実行結果を解析したレポート</p>
                <p>=== 実行条件 ============================================================</p>
                <p>start_time = {self.start_time.isoformat()}</p>
                <p>base_date_from = {base_date_from.isoformat()}</p>
                <p>base_date_to = {base_date_to.isoformat()}</p>
                <p>report_term = {self.stats_analysis_report_input.report_term}</p>
                <p>totalling_term = {self.stats_analysis_report_input.totalling_term}</p>
                <p>=========================================================================</p>
            </body>
        </html>'''
        # メール送信は一時停止
        #mail_attach_send(title=title, msg=msg, filepath=file_path)

    def collect_result_analysis_report_edit_header(
        self, ws: Worksheet,            # レポートの出力先のワークシート
        sheet_name: str,                # レポートの出力先のワークシートのシート名
        analysis_columns_info: list,):  # レポート出力対象カラムに関する情報リスト
        '''集計結果解析レポート用Excelの編集(見出し)'''

        ws.title = sheet_name  # シート名変更

        # 罫線定義
        side = Side(style='thin', color='000000')
        border = Border(top=side, bottom=side, left=side, right=side)
        # 背景色
        fill1 = PatternFill(patternType='solid', fgColor='0066CC')

        for i, col_info in enumerate(analysis_columns_info):
            # 見出し１行目
            head1_cell: Cell = ws[get_column_letter(i + 1) + str(1)]
            ws[head1_cell.coordinate] = col_info['head1']
            head1_cell.fill = fill1
            head1_cell.border = border
            head1_cell.alignment = Alignment(
                horizontal="center")  # 選択範囲内中央寄せ

    def collect_result_analysis_report_edit_body(
        self,
        ws: Worksheet,                  # レポートの出力先のワークシート
        spider_list: pd.Series,         # スパイダーの一覧
        result_df: pd.DataFrame,        # mongoDBへの集計結果より作成したデータフレーム
        check_col: str,                 # チェックを行いたいカラム
        analysis_columns_info: list,):  # レポート出力対象カラムに関する情報リスト
        '''集計結果解析レポート用Excelの編集'''

        # 罫線定義
        side = Side(style='thin', color='000000')
        border = Border(top=side, bottom=side, left=side, right=side)

        # 集計期間の数を確認
        aggregate_base_term_list_count = len(
            result_df['aggregate_base_term'].drop_duplicates())

        # エクセルシートの編集開始行数初期化
        base_row_idx: int = 2
        # スパイダー別に解析を行う。
        for spider in spider_list:
            columns_info_by_spider = deepcopy(
                analysis_columns_info)
            # スパイダー別のデータフレームを作成
            by_spider_df = result_df.query(
                f'spider_name == "{spider}"')
            # データ無しの件数を確認 ※スパイダーの可動前はデータ無しとなる。
            data_none_count: int = len(
                by_spider_df.query(f'{check_col} == ""'))
            nomal_count: int = aggregate_base_term_list_count - data_none_count
            # ステータス別の一覧を作成
            status_list_sr: pd.Series = (
                by_spider_df[check_col].drop_duplicates())

            # ステータス別にの出現件数が集計期間数と一致するか確認
            for status in status_list_sr:
                # データ無しは除く
                if status:
                    by_status_df = by_spider_df.query(
                        f'{check_col} == "{status}"')
                    if nomal_count != len(by_status_df):
                        for col_info in columns_info_by_spider:
                            if col_info['col'] == check_col:
                                col_info['warning_value'].append(status)

            # 列ごとにエクセルに編集
            for col_idx, col_info in enumerate(columns_info_by_spider):
                # データフレームの列ごとに１行ずつ処理を実施
                for row_idx, value in enumerate(by_spider_df[col_info['col']]):
                    # 更新対象のセルに値を設定
                    target_cell: Cell = ws[get_column_letter(
                        col_idx + 1) + str(base_row_idx + row_idx)]
                    ws[target_cell.coordinate] = value

                    # 同値カラー調整列の場合、更新対象セルの上と同値ならば色を変える。
                    if 'equivalent_color' in col_info:
                        compare_cell: Cell = ws[get_column_letter(
                            col_idx + 1) + str(base_row_idx + row_idx - 1)]
                        if target_cell.value == compare_cell.value:
                            target_cell.font = Font(
                                color=col_info['equivalent_color'])

                    # ワーニングカラー調整列の場合、ワーニングとする値をがあった場合、背景色を設定。
                    if 'warning_value' in col_info:
                        if value in col_info['warning_value']:
                            target_cell.fill = PatternFill(
                                patternType='solid', fgColor=col_info['warning_background_color'])

            # １つのスパイダーに対する編集が終わった段階で、次のスパイダー用にエクセルの編集行の基準を設定する。
            base_row_idx = base_row_idx + len(by_spider_df)

        # 各明細行のセルに罫線を設定する。
        max_cell: str = get_column_letter(
            ws.max_column) + str(ws.max_row)  # "BC55"のようなセル番地を生成
        cell: Cell  # 型ヒントの定義のみ
        for cells in ws[f'a2:{max_cell}']:
            for cell in cells:
                cell.border = border

        # 列ごとに次の処理を行う。
        # 最大幅を確認
        # それに合わせた幅を設定する。
        for cols in ws.columns:
            max_length = 0
            column = cols[0].column_letter  # 列名A,Bなどを取得
            for cell in cols:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            ws.column_dimensions[column].width = (max_length + 2.07)

        # ウィンドウ枠の固定。１行２列は常に表示させる。
        ws.freeze_panes = 'c2'

    def stats_analysis_report_edit_header(self, ws: Worksheet):
        '''レポート用Excelの見出し編集'''
        ws.title = 'Stats Analysis Report'  # シート名変更

        # 罫線定義
        side = Side(style='thin', color='000000')
        border = Border(top=side, bottom=side, left=side, right=side)
        fill1 = PatternFill(patternType='solid', fgColor='0066CC')
        fill2 = PatternFill(patternType='solid', fgColor='0099CC')

        for i, col_info in enumerate(self.stats_analysis_columns_info):
            # 見出し１行目
            head1_cell: Cell = ws[get_column_letter(i + 1) + str(1)]
            ws[head1_cell.coordinate] = col_info['head1']
            head1_cell.fill = fill1
            head1_cell.border = border
            head1_cell.alignment = Alignment(
                horizontal="centerContinuous")  # 選択範囲内中央寄せ

            # 見出し２行目
            head2_cell: Cell = ws[get_column_letter(i + 1) + str(2)]
            ws[head2_cell.coordinate] = col_info['head2']
            head2_cell.fill = fill2
            head2_cell.border = border
            head2_cell.alignment = Alignment(horizontal="center")  # 中央寄せ

    def stats_analysis_report_edit_body(self, ws: Worksheet, spider_result_all_df: pd.DataFrame):
        ''' '''

        '''以下のチェックを追加してみよう
        １．保存件数ゼロ
        ２．critical,errorあり
        ３．メモリ使用量の平均の1.5倍を超えているセル
        ４．リトライ件数が毎日ある場合OK。一部ある場合はワーニング。
        ５．処理時間(最大) > 600秒
        '''

        # 罫線定義
        side = Side(style='thin', color='000000')
        border = Border(top=side, bottom=side, left=side, right=side)

        # 列ごとにエクセルに編集
        for col_idx, col_info in enumerate(self.stats_analysis_columns_info):
            for row_idx, value in enumerate(spider_result_all_df[col_info['col']]):
                # 更新対象のセル
                target_cell: Cell = ws[get_column_letter(
                    col_idx + 1) + str(row_idx + 3)]

                # 表示単位の切り上げ (example:b -> kb)
                custom_value = value
                if 'digit_adjustment' in col_info:
                    if value:
                        custom_value = int(
                            value) // col_info['digit_adjustment']
                        Decimal(str(custom_value)).quantize(
                            Decimal('0'), rounding=ROUND_HALF_UP)

                # 小数点以下の調整
                if 'number_format' in col_info:
                    target_cell.number_format = col_info['number_format']
                else:
                    target_cell.number_format = '#,##0'

                # 更新対象のセルに値を設定
                ws[target_cell.coordinate] = custom_value

                # 同値カラー調整
                if 'equivalent_color' in col_info:
                    # 比較用の１つ上のセルと同じ値の場合は文字色を変更
                    compare_cell: Cell = ws[get_column_letter(
                        col_idx + 1) + str(row_idx + 2)]
                    if target_cell.value == compare_cell.value:
                        target_cell.font = Font(
                            color=col_info['equivalent_color'])

        # 一番右下のセル
        max_cell: str = get_column_letter(
            ws.max_column) + str(ws.max_row)  # "BC55"のようなセル番地を生成

        for cells in ws[f'a3:{max_cell}']:
            for cell in cells:
                cell: Cell
                cell.border = border

        # 列ごとに次の処理を行う。
        # 最大幅を確認
        # それに合わせた幅を設定する。
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # 列名A,Bなどを取得
            for cell in col:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            ws.column_dimensions[column].width = (max_length + 2.07)

        # ウィンドウ枠の固定。２行２列は常に表示させる。
        ws.freeze_panes = 'c3'
